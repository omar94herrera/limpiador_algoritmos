from tkinter import *
from openpyxl import *
import shutil, os
from datetime import datetime, timedelta
from random import randrange
import os.path
import logging

meses={1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}

fechas={"semana 1":((1,0,0,0),(7,23,59,59)),"semana 2":((8,0,0,0),(14,23,59,59)),"semana 3":((15,0,0,0),(21,23,59,59)),"semana 4":((22,0,0,0),(28,23,59,59)),"semana 5":((29,0,0,0),(31,23,59,59))}

febrero=["febrero","Febrero","FEBRERO"]

fotos=["FOTOS","Fotos","fotos"]

limites=[259200,604800]


def test():
    años=os.listdir(".")
    años.remove("ejecutable.py")
    años.remove("instrucciones.pdf")
    mesesalimpiar=mesesalimpiarV.get()
    mesesalimpiar=mesesalimpiar.split(";")
    print(mesesalimpiar)
    aps={}
    for i in mesesalimpiar:
        i=i.split()
        try:
            aps[i[0]]=i[1].split("-")
        except:
            pass
    print(aps)
    for año in años:
        for mes in os.listdir(".\\"+año):
            if año in aps:
                if mes not in aps[año]:
                    continue
                else:
                    print(año,mes)
            else:
                continue


def orden66():
    años=os.listdir(".")
    try:
        crear_carpetas_brutas()
        crear_carpetas_limpias()
    except:
        pass
    años.remove("ejecutable.py")
    años.remove("instrucciones.pdf")
    mesesalimpiar=mesesalimpiarV.get()
    mesesalimpiar=mesesalimpiar.split(";")
    semanasalimpiar=semanasalimpiarV.get()
    semanasalimpiar=semanasalimpiar.split(",")
    ubicacionesalimpiar=ubicacionesalimpiarV.get()
    ubicacionesalimpiar=ubicacionesalimpiar.split()#ubicaciones van separadas por espacio
    maquinasalimpiar=maquinasalimpiarV.get()
    maquinasalimpiar=maquinasalimpiar.split()#maquinas van separadas por espacio
    print(maquinasalimpiar)
    print(ubicacionesalimpiar)
    print(mesesalimpiar)
    print(semanasalimpiar)
    aps={}
    for i in mesesalimpiar:
        i=i.split()
        aps[i[0]]=i[1].split("-")
    for año in años:
        for mes in os.listdir(".\\"+año):
            if año in aps:
                if mes not in aps[año]:
                    continue
            else:
                continue
            numeromes=0
            for i,j in meses.items():
                if j==mes:
                    numeromes=i
            for fechitas,tuplas in fechas.items():
                if fechitas not in semanasalimpiar and semanasalimpiar!=['']:
                    continue
                if mes in febrero and fechitas=="semana 5":
                    continue
                if fechitas!="semana 5":
                    fecha_inicio=datetime(int(año),numeromes,tuplas[0][0],tuplas[0][1],tuplas[0][2],tuplas[0][3])
                    fecha_termino=datetime(int(año),numeromes,tuplas[1][0],tuplas[1][1],tuplas[1][2],tuplas[1][3])
                else:
                    if mes not in febrero:
                        try:
                            fecha_inicio=datetime(int(año),numeromes,tuplas[0][0],tuplas[0][1],tuplas[0][2],tuplas[0][3])
                            fecha_termino=datetime(int(año),numeromes,tuplas[1][0],tuplas[1][1],tuplas[1][2],tuplas[1][3])
                        except:
                            fecha_inicio=datetime(int(año),numeromes,tuplas[0][0],tuplas[0][1],tuplas[0][2],tuplas[0][3])
                            fecha_termino=datetime(int(año),numeromes,30,tuplas[1][1],tuplas[1][2],tuplas[1][3])
                lugares_maquinas={}
                n=semana_del_mes(fecha_inicio,fecha_termino)
                for semana in os.listdir(".\\"+año+"\\"+mes):
                    ruido=semana.split()
                    fecha=ruido[2].split("-")
                    ruido="Ruido tronaduras "+ruido[2]
                    for ubicacion in os.listdir(".\\"+año+"\\"+mes+"\\"+semana+"\\"+ruido):
                        if ubicacion not in fotos:
                            for archivo in os.listdir(".\\"+año+"\\"+mes+"\\"+semana+"\\"+ruido+"\\"+ubicacion):
                                excel=archivo.split(".")
                                numero=len(excel)-1
                                tipo=excel[numero]
                                try:
                                    nombre=excel[0].split("-")
                                    lugari=nombre[0]
                                    maquinai=nombre[1]
                                    if tipo in ["xls","txt"]:
                                        try:
                                            lugares_maquinas[lugari].add(maquinai)
                                        except:
                                            lugares_maquinas[lugari]={maquinai}
                                except:
                                    pass
                for lugar,maquinas in lugares_maquinas.items():
                    if lugar not in ubicacionesalimpiar and ubicacionesalimpiar!=[]:
                        continue
                    for maquina in maquinas:
                        if maquina not in maquinasalimpiar and maquinasalimpiar!=[]:
                            continue
                        try: #en caso de que el archivo no exista
                            print("se esta analizando el archivo", lugar+"-"+maquina,"de", mes, "del año",año)
                            numeromagico=crear_archivos_brutos(lugar,maquina,fecha_inicio,fecha_termino,mes,año)
                            nombre_archivo_bruto="semana "+n+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)+".xlsx"
                            limpiar_archivo(nombre_archivo_bruto,mes,numeromagico,fecha_inicio,fecha_termino)
                            nombre_archivo_limpio=mes+"-"+lugar+"-"+maquina+"-semana "+n+".xlsx"
                            shutil.move(".\\"+nombre_archivo_bruto,".\\archivos brutos\\"+año+"\\"+mes+"\\semana "+n)
                            print("se movio el archivo bruto")
                            shutil.move(".\\"+nombre_archivo_limpio,".\\archivos limpios\\"+año+"\\"+mes+"\\semana "+n)
                            print("se movio el archivo limpio")
                        except:
                            logging.exception("message")
                            #pass
                del(lugares_maquinas)
    print("todo listo")
    ventana_fin()
                    


#las fechas de inicio y termino las ingreso en formato fecha
def crear_archivos_brutos(lugar,maquina,fecha_inicio,fecha_termino,mes,año): 
    n=int(semana_del_mes(fecha_inicio,fecha_termino))
    archivoxls=lugar+"-"+maquina+".xls"
    archivotxt=lugar+"-"+maquina+".txt"

    if n==1:
        if mes!="Enero":
            nm=[]
            try: #el try es en caso de que no exista una carpeta del mes anterior
                mesant=meses[fecha_inicio.month-1] 
                fechas=[]
                for semana in os.listdir(".\\"+año+"\\"+mesant):
                    semana=semana.split()
                    fecha=semana[2].split("-")
                    fechas.append(datetime(int(fecha[2]),int(fecha[1]),int(fecha[0])))
                    fechas.sort()
                diaa=""
                mess=""
                ultima_semana=len(fechas)-1
                if len(str(fechas[ultima_semana].day))==1:
                    diaa="0"+str(fechas[ultima_semana].day)
                else:
                    diaa=str(fechas[ultima_semana].day)
                if len(str(fechas[ultima_semana].month))==1:
                    mess="0"+str(fechas[ultima_semana].month)
                else:
                    mess=str(fechas[ultima_semana].month)
                fechaaaa=diaa+"-"+mess+"-"+str(fechas[ultima_semana].year)
                ruta=".\\"+año+"\\"+mesant+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
                nm.append(nmi)
            except: 
                pass

            fechasdespuesfechainicio=[]
            fechasdespuesfechatermino=[]
            fechas=[]
            for semana in os.listdir(".\\"+año+"\\"+mes): #guardo las fechas de las carpetas semana en formato fecha
                semana=semana.split()
                fecha=semana[2].split("-")
                fechita=datetime(int(fecha[2]),int(fecha[1]),int(fecha[0]))
                if fechita>=fecha_inicio:
                    fechasdespuesfechainicio.append(fechita)
                if fechita>=fecha_termino:
                    fechasdespuesfechatermino.append(fechita)
                fechasdespuesfechainicio.sort()
                fechasdespuesfechatermino.sort()
            if fechasdespuesfechainicio!=[] and fechasdespuesfechatermino!=[]:
                if fechasdespuesfechainicio[0]!=fechasdespuesfechatermino[0]:
                    fechas.append(fechasdespuesfechainicio[0])
                    fechas.append(fechasdespuesfechatermino[0])
                    
                if fechasdespuesfechainicio[0]==fechasdespuesfechatermino[0]:
                    fechas.append(fechasdespuesfechainicio[0])
          
                    
            elif fechasdespuesfechainicio==[] and fechasdespuesfechatermino!=[]:
                fechas.append(fechasdespuesfechatermino[0])
            elif fechasdespuesfechainicio!=[] and fechasdespuesfechatermino==[]:
                fechas.append(fechasdespuesfechainicio[0])

                
            for i in fechas:
                diaa=""
                mess=""
                if len(str(i.day))==1:
                    diaa="0"+str(i.day)
                else:
                    diaa=str(i.day)
                if len(str(i.month))==1:
                    mess="0"+str(i.month)
                else:
                    mess=str(i.month)
                fechaaaa=diaa+"-"+mess+"-"+str(i.year)
                ruta=".\\"+año+"\\"+mes+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
            nm.append(nmi)
            
            
        else:
            nm=[]
            try: #el try es en caso de que no exista el mes de diciembre en el año anterior
                añoant=str(int(año)-1)
                mesant="Diciembre"
                fechas=[]
                for semana in os.listdir(".\\"+añoant+"\\"+mesant):
                    semana=semana.split()
                    fecha=semana[2].split("-")
                    fechas.append(datetime(int(fecha[2]),int(fecha[1]),int(fecha[0])))
                    fechas.sort()
                diaa=""
                mess=""
                ultima_semana=len(fechas)-1
                if len(str(fechas[ultima_semana].day))==1:
                    diaa="0"+str(fechas[ultima_semana].day)
                else:
                    diaa=str(fechas[ultima_semana].day)
                if len(str(fechas[ultima_semana].month))==1:
                    mess="0"+str(fechas[ultima_semana].month)
                else:
                    mess=str(fechas[ultima_semana].month)
                fechaaaa=diaa+"-"+mess+"-"+str(fechas[ultima_semana].year)
                ruta=".\\"+añoant+"\\"+mesant+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
                nm.append(nmi)
            except:
                pass
            
            fechasdespuesfechainicio=[]
            fechasdespuesfechatermino=[]
            fechas=[]
            for semana in os.listdir(".\\"+año+"\\"+mes): #guardo las fechas de las carpetas semana en formato fecha
                semana=semana.split()
                fecha=semana[2].split("-")
                fechita=datetime(int(fecha[2]),int(fecha[1]),int(fecha[0]))
                if fechita>=fecha_inicio:
                    fechasdespuesfechainicio.append(fechita)
                if fechita>=fecha_termino:
                    fechasdespuesfechatermino.append(fechita)
                fechasdespuesfechainicio.sort()
                fechasdespuesfechatermino.sort()
            if fechasdespuesfechainicio!=[] and fechasdespuesfechatermino!=[]:
                if fechasdespuesfechainicio[0]!=fechasdespuesfechatermino[0]:
                    fechas.append(fechasdespuesfechainicio[0])
                    fechas.append(fechasdespuesfechatermino[0])
                    #print("ambas listas no vacias y distintas")
                if fechasdespuesfechainicio[0]==fechasdespuesfechatermino[0]:
                    fechas.append(fechasdespuesfechainicio[0])
                    #print("ambas listas no vacias e iguales en su primer elemento")
                    
            elif fechasdespuesfechainicio==[] and fechasdespuesfechatermino!=[]:
                fechas.append(fechasdespuesfechatermino[0])
                #print("fechadespuesinicio vacia y la otra no")
            elif fechasdespuesfechainicio!=[] and fechasdespuesfechatermino==[]:
                fechas.append(fechasdespuesfechainicio[0])
                #print("fechadespuestermino vacia y la otra no")

                
            for i in fechas:
                diaa=""
                mess=""
                if len(str(i.day))==1:
                    diaa="0"+str(i.day)
                else:
                    diaa=str(i.day)
                if len(str(i.month))==1:
                    mess="0"+str(i.month)
                else:
                    mess=str(i.month)
                fechaaaa=diaa+"-"+mess+"-"+str(i.year)
                ruta=".\\"+año+"\\"+mes+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
            nm.append(nmi)
            



    if n>1 and n<=4:
        fechasdespuesfechainicio=[]
        fechasdespuesfechatermino=[]
        fechasantesfechainicio=[]
        fechas=[]
        for semana in os.listdir(".\\"+año+"\\"+mes): #guardo las fechas de las carpetas semana en formato fecha
            semana=semana.split()
            fecha=semana[2].split("-")
            fechita=datetime(int(fecha[2]),int(fecha[1]),int(fecha[0]))
            if fechita<=fecha_inicio:
                fechasantesfechainicio.append(fechita)
            if fechita>=fecha_inicio:
                fechasdespuesfechainicio.append(fechita)
            if fechita>=fecha_termino:
                fechasdespuesfechatermino.append(fechita)
        fechasantesfechainicio.sort()
        fechasdespuesfechainicio.sort()
        fechasdespuesfechatermino.sort()
        if fechasantesfechainicio!=[]:
            nani=len(fechasantesfechainicio)-1
            fechas.append(fechasantesfechainicio[nani])
        if fechasdespuesfechainicio!=[] and fechasdespuesfechatermino!=[]:
            if fechasdespuesfechainicio[0]!=fechasdespuesfechatermino[0]:
                fechas.append(fechasdespuesfechainicio[0])
                fechas.append(fechasdespuesfechatermino[0])
            if fechasdespuesfechainicio[0]==fechasdespuesfechatermino[0]:
                fechas.append(fechasdespuesfechainicio[0])                 
        elif fechasdespuesfechainicio==[] and fechasdespuesfechatermino!=[]:
            fechas.append(fechasdespuesfechatermino[0])

        elif fechasdespuesfechainicio!=[] and fechasdespuesfechatermino==[]:
            fechas.append(fechasdespuesfechainicio[0])
        fechas.sort()
        
        nm=[]
        for i in fechas:
            diaa=""
            mess=""
            if len(str(i.day))==1:
                diaa="0"+str(i.day)
            else:
                diaa=str(i.day)
            if len(str(i.month))==1:
                mess="0"+str(i.month)
            else:
                mess=str(i.month)
            fechaaaa=diaa+"-"+mess+"-"+str(i.year)
            ruta=".\\"+año+"\\"+mes+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
            if os.path.exists(ruta+"\\"+archivoxls):
                nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
            elif os.path.exists(ruta+"\\"+archivotxt):
                nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
            nm.append(nmi)
        if len(nm)==3:
            nm.remove(nm[2])
        
            
        


    elif n==5:
        fechasdespuesfechainicio=[]
        fechasdespuesfechatermino=[]
        fechasantesfechainicio=[]
        fechas=[]
        for semana in os.listdir(".\\"+año+"\\"+mes): #guardo las fechas de las carpetas semana en formato fecha
            semana=semana.split()
            fecha=semana[2].split("-")
            fechita=datetime(int(fecha[2]),int(fecha[1]),int(fecha[0]))
            if fechita<=fecha_inicio:
                fechasantesfechainicio.append(fechita)
            if fechita>=fecha_inicio:
                fechasdespuesfechainicio.append(fechita)
            if fechita>=fecha_termino:
                fechasdespuesfechatermino.append(fechita)
        fechasantesfechainicio.sort()
        fechasdespuesfechainicio.sort()
        fechasdespuesfechatermino.sort()
        if fechasantesfechainicio!=[]:
            nani=len(fechasantesfechainicio)-1
            fechas.append(fechasantesfechainicio[nani])
        if fechasdespuesfechainicio!=[] and fechasdespuesfechatermino!=[]:
            if fechasdespuesfechainicio[0]!=fechasdespuesfechatermino[0]:
                fechas.append(fechasdespuesfechainicio[0])
                fechas.append(fechasdespuesfechatermino[0])

            if fechasdespuesfechainicio[0]==fechasdespuesfechatermino[0]:
                fechas.append(fechasdespuesfechainicio[0])                 
        elif fechasdespuesfechainicio==[] and fechasdespuesfechatermino!=[]:
            fechas.append(fechasdespuesfechatermino[0])
        elif fechasdespuesfechainicio!=[] and fechasdespuesfechatermino==[]:
            fechas.append(fechasdespuesfechainicio[0])
        fechas.sort()
        nm=[]
           
        for i in fechas:
            diaa=""
            mess=""
            if len(str(i.day))==1:
                diaa="0"+str(i.day)
            else:
                diaa=str(i.day)
            if len(str(i.month))==1:
                mess="0"+str(i.month)
            else:
                mess=str(i.month)
            fechaaaa=diaa+"-"+mess+"-"+str(i.year)
            ruta=".\\"+año+"\\"+mes+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
            if os.path.exists(ruta+"\\"+archivoxls):
                nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
            elif os.path.exists(ruta+"\\"+archivotxt):
                nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
            nm.append(nmi)

        if len(nm)==3:
            nm.remove(nm[2])
        
        
        if mes!="Diciembre":
            try: #el try es en caso de que no exista una carpeta con el mes siguiente 
                messgte=meses[fecha_inicio.month+1]
                fechas=[]
                for semana in os.listdir(".\\"+año+"\\"+messgte):
                    semana=semana.split()
                    fecha=semana[2].split("-")
                    fechas.append(datetime(int(fecha[2]),int(fecha[1]),int(fecha[0])))
                    fechas.sort()
                    
                diaa=""
                mess=""
                nm=[]
                if len(str(fechas[0].day))==1:
                    diaa="0"+str(fechas[0].day)
                else:
                    diaa=str(fechas[0].day)
                if len(str(fechas[0].month))==1:
                    mess="0"+str(fechas[0].month)
                else:
                    mess=str(fechas[0].month)
                fechaaaa=diaa+"-"+mess+"-"+str(fechas[0].year)
                ruta=".\\"+año+"\\"+messgte+"\\Marcha Blanca "+fechaaaa+"\\Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
                nm.append(nmi)
            except:
                #logging.exception("message")
                pass
        else:
            try:#el try es en caso de que no exista una carpeta con el mes de enero en el año siguiente
                messgte="Enero" 
                añosgte=str(int(año)+1) #el año se ingresa como string
                fechas=[]
                for semana in os.listdir(".\\"+añosgte+"\\"+messgte):
                    semana=semana.split()
                    fecha=semana[2].split("-")
                    fechas.append(datetime(int(fecha[2]),int(fecha[1]),int(fecha[0])))
                fechas.sort()
                diaa=""
                mess=""
                if len(str(fechas[0].day))==1:
                    diaa="0"+str(fechas[0].day)
                else:
                    diaa=str(fechas[0].day)
                if len(str(fechas[0].month))==1:
                    mess="0"+str(fechas[0].month)
                else:
                    mess=str(fechas[0].month)
                fechaaaa=diaa+"-"+mess+"-"+str(fechas[0].year)
                ruta=".\\"+añosgte+"\\"+messgte+"\\"+"Marcha Blanca "+fechaaaa+"\\"+"Ruido tronaduras "+fechaaaa+"\\"+lugar
                if os.path.exists(ruta+"\\"+archivoxls):
                    nmi=copiar_xls(fecha_inicio,fecha_termino,archivoxls,ruta)
                elif os.path.exists(ruta+"\\"+archivotxt):
                    nmi=copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta)
                nm.append(nmi)

                if len(nm)==3:
                    nm.remove(nm[2])
                
    
            except:
                #logging.exception("message")
                pass
    
    return(nm)

def copiar_xls(fecha_inicio,fecha_termino,archivo,ruta):
    aaaa=archivo.split(".")
    lugar=aaaa[0].split("-")[0]
    maquina=aaaa[0].split("-")[1]
    x=semana_del_mes(fecha_inicio,fecha_termino)
    año=str(fecha_inicio.year)
    
    ruta1 = ruta + os.sep
    origen = ruta1 + archivo
    destino = ruta1 + 'copia.txt'
    try:
        archivo = shutil.copy2(origen, destino)
    except:
        print('Error en la copia del archivo', archivo, "del año", año, "no existe" )
 
    try: #aqui si el archivo existe lo abro,sino hago uno nuevo
        rutaa=".\\semana "+x+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)+".xlsx"
        libro=load_workbook(rutaa) 
            
        hojas=libro.sheetnames
        hoja=libro[hojas[0]]
        #verificador=0
        archivo=open(destino)
        nlinea=1
        nfilas=0
        numeromagico=hoja.max_row
        if numeromagico==1:
            numeromagico=0
        #print(numeromagico)
        veri=0
        if numeromagico not in limites:
            for linea in archivo:
                if nlinea>3:
                    linea=linea.split()
                    fechaa=linea[0]
                    horaa=linea[1]
                    año=int(linea[0].split("/")[0])
                    mes=int(linea[0].split("/")[1])
                    dia=int(linea[0].split("/")[2])
                    hora=int(linea[1].split(":")[0])
                    minuto=int(linea[1].split(":")[1])
                    segundo=""
                    contadorS=0
                    for i in linea[1].split(":")[2]:
                        if contadorS<2:
                            segundo+=i
                        contadorS+=1
                    segundo=int(segundo)
                    fecha=datetime(año,mes,dia,hora,minuto,segundo)
                    if numeromagico!=0 and nfilas==0 and veri==0:
                        veri=1
                        asd=hoja["A"+str(numeromagico)].value
                        #print("blablabla",asd)
                        if asd!=None:
                            asd=asd.split()
                            año=int(asd[0].split("/")[0])
                            mes=int(asd[0].split("/")[1])
                            dia=int(asd[0].split("/")[2])
                            hora=int(asd[2].split(":")[0])
                            minuto=int(asd[2].split(":")[1])
                            segundo=""
                            contadorS=0
                            for i in asd[2].split(":")[2]:
                                if contadorS<2:
                                    segundo+=i
                                contadorS+=1
                            segundo=int(segundo)
                            fecha2=datetime(año,mes,dia,hora,minuto,segundo)
                            #print(fecha2)
                            delta=fecha-fecha2
                            dif=delta.days*3600*24+delta.seconds
                            numeromagico+=dif-1
                            #print("yolo",dif,numeromagico)

                    try:   
                        if fecha>=fecha_inicio and fecha<=fecha_termino:
                            fechota=linea[0]+" - "+linea[1]
                            nfilas+=1
                            if hoja["A1"].value==None:
                                hoja["A1"]=fechota
                                hoja["B1"]=linea[2]
                                hoja["C1"]=linea[3]
                                hoja["D1"]=linea[4]
                            else:
                                hoja.append([fechota,linea[2],linea[3],linea[4]])
                    except:
                        pass
                          
                nlinea+=1
        try:
            #print(fecha)
            if fecha!=fecha_termino:
                numeromagico+=nfilas
                nombre_archivo="semana "+n+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)
                print("ver fila", numeromagico,"en archivo", nombre_archivo,"en", meses[fecha_inicio.month])
        except:
            pass
        libro.save(rutaa)
        archivo.close()
        #if verificador==1:
            #continue #ver si hago algo aca, demas que sale algo util
        os.remove(destino)
        return(numeromagico)

    except:
        #logging.exception("message")
        libro=Workbook()
        hojas=libro.sheetnames
        hoja=libro[hojas[0]]
        #verificador=0
        archivo=open(destino)
        nlinea=1
        nfilas=1
        for linea in archivo:
            if nlinea>3 :
                linea=linea.strip().split()
                fechaa=linea[0]
                horaa=linea[1]
                año=int(linea[0].split("/")[0])
                mes=int(linea[0].split("/")[1])
                dia=int(linea[0].split("/")[2])
                hora=int(linea[1].split(":")[0])
                minuto=int(linea[1].split(":")[1])
                segundo=""
                contadorS=0
                for i in linea[1].split(":")[2]:
                    if contadorS<2:
                        segundo+=i
                    contadorS+=1
                segundo=int(segundo)
                fecha=datetime(año,mes,dia,hora,minuto,segundo)



                    
                if fecha>=fecha_inicio and fecha<=fecha_termino:
                    #verificador=1
                    posfecha="A"+str(nfilas)
                    posdato1="B"+str(nfilas)
                    posdato2="C"+str(nfilas)
                    posdato3="D"+str(nfilas)
                    fechita=linea[0]+" - "+linea[1]
                    hoja[posfecha]=fechita
                    hoja[posdato1]=linea[2]
                    hoja[posdato2]=linea[3]
                    hoja[posdato3]=linea[4]
                    nfilas+=1
                        
                        
            nlinea=nlinea+1
        archivo.close()
        #if verificador==1:
            #continue #ver si hago algo aca, demas que sale algo util
            #podria ser que si el verificador es 0, el archivo se borre
            #o jamas se cree
        os.remove(destino)
        hoja.column_dimensions["A"].width=23
        #recordar que x es el numero de semana como string
                
        nombre_archivo="semana "+x+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)
        libro.save(filename=nombre_archivo+".xlsx")
        if nfilas-1!=0:
            print("ver fila", nfilas-1,"en archivo", nombre_archivo,"en", meses[mes])
        return(nfilas-1)



def copiar_txt(fecha_inicio,fecha_termino,archivotxt,ruta):
    aaaa=archivotxt.split(".")
    lugar=aaaa[0].split("-")[0]
    maquina=aaaa[0].split("-")[1]
    x=semana_del_mes(fecha_inicio,fecha_termino)
    año=str(fecha_inicio.year)
 
    try: #aqui si el archivo existe lo abro,sino hago uno nuevo
        rutaa=".\\semana "+n+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)+".xlsx"
        libro=load_workbook(rutaa) 
            
        hojas=libro.sheetnames
        hoja=libro[hojas[0]]
        #verificador=0
        archivo=open(rutaa)
        nlinea=1
        nfilas=0
        numeromagico=hoja.max_row
        if numeromagico==1:
            numeromagico=0
        #print(numeromagico)
        veri=0
        if numeromagico not in limites:
            for linea in archivo:
                if nlinea>3:
                    linea=linea.split()
                    fechaa=linea[0]
                    horaa=linea[1]
                    año=int(linea[0].split("/")[0])
                    mes=int(linea[0].split("/")[1])
                    dia=int(linea[0].split("/")[2])
                    hora=int(linea[1].split(":")[0])
                    minuto=int(linea[1].split(":")[1])
                    segundo=""
                    contadorS=0
                    for i in linea[1].split(":")[2]:
                        if contadorS<2:
                            segundo+=i
                        contadorS+=1
                    segundo=int(segundo)
                    fecha=datetime(año,mes,dia,hora,minuto,segundo)
                    if numeromagico!=0 and nfilas==0 and veri==0:
                        veri=1
                        asd=hoja["A"+str(numeromagico)].value
                        #print("blablabla",asd)
                        if asd!=None:
                            asd=asd.split()
                            año=int(asd[0].split("/")[0])
                            mes=int(asd[0].split("/")[1])
                            dia=int(asd[0].split("/")[2])
                            hora=int(asd[2].split(":")[0])
                            minuto=int(asd[2].split(":")[1])
                            segundo=""
                            contadorS=0
                            for i in asd[2].split(":")[2]:
                                if contadorS<2:
                                    segundo+=i
                                contadorS+=1
                            segundo=int(segundo)
                            fecha2=datetime(año,mes,dia,hora,minuto,segundo)
                            #print(fecha2)
                            delta=fecha-fecha2
                            dif=delta.days*3600*24+delta.seconds
                            numeromagico+=dif-1
                            #print("yolo",dif,numeromagico)

                    try:  
                        if fecha>=fecha_inicio and fecha<=fecha_termino:
                            fechota=linea[0]+" - "+linea[1]
                            nfilas+=1
                            if hoja["A1"].value==None:
                                hoja["A1"]=fechota
                                hoja["B1"]=linea[2]
                                hoja["C1"]=linea[3]
                                hoja["D1"]=linea[4]
                            else:
                                hoja.append([fechota,linea[2],linea[3],linea[4]])
                    except:
                        pass
                          
                nlinea+=1
        try:
            #print(fecha)
            if fecha!=fecha_termino:
                numeromagico+=nfilas
                nombre_archivo="semana"+" "+ x +" "+lugar+" "+maquina
                print("ver fila", numeromagico,"en archivo", nombre_archivo,"en", meses[fecha_inicio.month])
        except:
            pass
        libro.save(rutaa)
        archivo.close()
        #if verificador==1:
            #pass #ver si hago algo aca, demas que sale algo util
        os.remove(destino)
        return(numeromagico)

    except:
        #logging.exception("message")
        libro=Workbook()
        hojas=libro.sheetnames
        hoja=libro[hojas[0]]
        #verificador=0
        archivo=open(ruta+"\\"+archivotxt)
        nlinea=1
        nfilas=1
        for linea in archivo:
            if nlinea>3 :
                linea=linea.split()
                fechaa=linea[0]
                horaa=linea[1]
                año=int(linea[0].split("/")[0])
                mes=int(linea[0].split("/")[1])
                dia=int(linea[0].split("/")[2])
                hora=int(linea[1].split(":")[0])
                minuto=int(linea[1].split(":")[1])
                segundo=""
                contadorS=0
                for i in linea[1].split(":")[2]:
                    if contadorS<2:
                        segundo+=i
                    contadorS+=1
                segundo=int(segundo)
                fecha=datetime(año,mes,dia,hora,minuto,segundo)

                    
                if fecha>=fecha_inicio and fecha<=fecha_termino:
                    #verificador=1
                    posfecha="A"+str(nfilas)
                    posdato1="B"+str(nfilas)
                    posdato2="C"+str(nfilas)
                    posdato3="D"+str(nfilas)
                    fechita=linea[0]+" - "+linea[1]
                    hoja[posfecha]=fechita
                    hoja[posdato1]=linea[2]
                    hoja[posdato2]=linea[3]
                    hoja[posdato3]=linea[4]
                    nfilas+=1
                        
                        
            nlinea=nlinea+1
        archivo.close()
        #if verificador==1:
            #continue #ver si hago algo aca, demas que sale algo util
            #podria ser que si el verificador es 0, el archivo se borre
            #o jamas se cree
        os.remove(destino)
        hoja.column_dimensions["A"].width=23
        #recordar que x es el numero de semana como string
                
        nombre_archivo="semana "+n+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)
        libro.save(filename=nombre_archivo+".xlsx")
        if nfilas-1!=0:
            print("ver fila", nfilas-1,"en archivo", nombre_archivo,"en", meses[mes])
        return(nfilas-1)









def limpiar_archivo(archivo,mes,numeromagico,fecha_inicio,fecha_termino):
    doc=load_workbook(archivo)
    H=doc.sheetnames
    hoja=doc[str(H[0])]
    nlinea=1
    fechas=[]
    datos=[]
    #verificador=0
    limsup=hoja.max_row
    #print(numeromagico,limsup)

    try:
        inicio=hoja["A1"].value
        inicio=inicio.split()
        fecha=inicio[0]
        fechaa=fecha.split("/")
        año=int(fechaa[0])
        mess=int(fechaa[1])
        dia=int(fechaa[2])
        hora=inicio[2]
        hora=hora.split(":")
        segundos=hora[2]
        a=""
        contadorS=0
        for i in segundos:
            if contadorS<2:
                a+=i
            contadorS+=1
        fechotainicio=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
        if fechotainicio-fecha_inicio>timedelta(0):
                di=fechotainicio-fecha_inicio
                print("faltan", di.days, "días con", di.seconds, "segundos al inicio del archivo", archivo)


        fin=hoja["A"+str(limsup)].value
        fin=fin.split()
        fecha=fin[0]
        fechaa=fecha.split("/")
        año=int(fechaa[0])
        mess=int(fechaa[1])
        dia=int(fechaa[2])
        hora=fin[2]
        hora=hora.split(":")
        segundos=hora[2]
        a=""
        contadorS=0
        for i in segundos:
            if contadorS<2:
                a+=i
            contadorS+=1
        fechotafin=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
        if fecha_termino-fechotafin>timedelta(0):
                dt=fecha_termino-fechotafin
                print("faltan", dt.days, "días con", dt.seconds, "segundos al final del archivo", archivo)
    except:
        #logging.exception("message")
        pass

    if len(numeromagico)==1 or 0 in numeromagico or len({i for i in numeromagico})==1 or set(numeromagico) & set(limites)!=set():
        comodin=[]
        for i in numeromagico:
            if i!=0:
                comodin.append(i)
        numeromagico=comodin
        #print(numeromagico,limsup)
        if numeromagico[0] not in limites and numeromagico[0] not in [i+1 for i in limites] :
            for i in hoja.iter_rows():
                for r,row in enumerate(i):
                    if row.value!=None: 
                        if r==0:
                            celda=row.value
                            celda=celda.split()
                            fecha=celda[0]
                            fechaa=fecha.split("/")
                            año=int(fechaa[0])
                            mess=int(fechaa[1])
                            dia=int(fechaa[2])
                            hora=celda[2]
                            hora=hora.split(":")
                            segundos=hora[2]
                            a=""
                            contadorS=0
                            for i in segundos:
                                if contadorS<2:
                                    a+=i
                                contadorS+=1
                            if nlinea==numeromagico[0]:
                                fechita1=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                            elif nlinea==numeromagico[0]+1:
                                fechita2=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                            hora=hora[0]+":"+hora[1]+":"+a
                            fechas.append(fecha+" - "+hora)
                        elif r==2:
                            datos.append(round(float(row.value),1))
                nlinea=nlinea+1       
            delta=fechita2-fechita1
            dif=delta.days*60*60*24+delta.seconds
            if dif>=60*60*24:
                print("faltan más de 24 horas de datos")
            tituloo=archivo.split()
            maquina=tituloo[3].split(".")[0]
            titulo=mes+"-"+tituloo[2]+"-"+maquina+"-"+tituloo[0]+" "+tituloo[1]
            libro=Workbook()
            hojas=libro.sheetnames
            hojaa=libro[hojas[0]]
            if numeromagico[0]!=hoja.max_row:
                for i in range(len(datos)):
                    k=i+1
                    if k<numeromagico[0]+1:
                        posfecha="A"+str(k)
                        posdatos="B"+str(k)
                        hojaa[posfecha]=fechas[i]
                        hojaa[posdatos]=datos[i]
                    if k==numeromagico[0]+1:
                        fechitaa=fechita1+timedelta(0,1)
                        for j in range(dif):
                            posfecha="A"+str(k+j)
                            posdatos="B"+str(k+j)
                            hojaa[posdatos]=randrange(40,70)########### Aquí se cambia aleatoriedad
                            fechita=str(fechitaa)
                            fechita=fechita.split()
                            hojaa[posfecha]= fechita[0]+" - "+fechita[1]
                            fechitaa+=timedelta(0,1)
                        fechitaa-=timedelta(0,1)
                        hojaa["B"+str(i+dif)]=datos[i]
                        hojaa["A"+str(i+dif)]=str(fechitaa).split()[0]+" - "+str(fechitaa).split()[1]
                    if k>numeromagico[0]+1:
                        posfecha="A"+str(i+dif)
                        posdatos="B"+str(i+dif)
                        hojaa[posfecha]=fechas[i]
                        hojaa[posdatos]=datos[i]

        else:
            for i in hoja.iter_rows():
                for r,row in enumerate(i):
                    if row.value!=None: 
                        if r==0:
                            celda=row.value
                            celda=celda.split()
                            fecha=celda[0]
                            fechaa=fecha.split("/")
                            año=int(fechaa[0])
                            mess=int(fechaa[1])
                            dia=int(fechaa[2])
                            hora=celda[2]
                            hora=hora.split(":")
                            segundos=hora[2]
                            a=""
                            contadorS=0
                            for i in segundos:
                                if contadorS<2:
                                    a+=i
                                contadorS+=1
                            hora=hora[0]+":"+hora[1]+":"+a
                            fechas.append(fecha+" - "+hora)
                        elif r==2:
                            datos.append(round(float(row.value),1))
                nlinea+=1
            tituloo=archivo.split()
            maquina=tituloo[3].split(".")[0]
            titulo=mes+"-"+tituloo[2]+"-"+maquina+"-"+tituloo[0]+" "+tituloo[1]
            libro=Workbook()
            hojas=libro.sheetnames
            hojaa=libro[hojas[0]]
            for i in range(len(datos)):
                posfecha="A"+str(i+1)
                posdatos="B"+str(i+1)
                hojaa[posfecha]=fechas[i]
                hojaa[posdatos]=datos[i]

        hojaa.column_dimensions["A"].width=23
            
        nombre_archivo=titulo+".xlsx"
        libro.save(filename=nombre_archivo)

    elif len({i for i in numeromagico})>1 and 0 not in numeromagico and 1 not in numeromagico:
        if limsup not in numeromagico:
            x=0
            y=0
            nlinea=1
            fechita1=0
            fechita2=0
            for i in hoja.iter_rows():
                for r,row in enumerate(i):
                    if row.value!=None: 
                        if r==0:
                            celda=row.value
                            celda=celda.split()
                            fecha=celda[0]
                            fechaa=fecha.split("/")
                            año=int(fechaa[0])
                            mess=int(fechaa[1])
                            dia=int(fechaa[2])
                            hora=celda[2]
                            hora=hora.split(":")
                            segundos=hora[2]
                            a=""
                            contadorS=0
                            for i in segundos:
                                if contadorS<2:
                                    a+=i
                                contadorS+=1
                            if nlinea in numeromagico and x==0:
                                x=nlinea
                                fechita1=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                                #print("x:", x," y fechita1:",fechita1)
                            elif nlinea in [i+1 for i in numeromagico] and x!=0 and y==0:
                                fechita2=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                                #print("fechita2:",fechita2)
                            if fechita1!=0 and fechita2!=0:
                                delta=fechita2-fechita1
                                dif1=delta.days*60*60*24+delta.seconds
                                if nlinea in [i+1-dif1 for i in numeromagico] and x!=0 and y==0:
                                    y=nlinea
                                    fechita3=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                                    #print("y:", y, " y fechita3:",fechita3)
                                elif nlinea in [i+2-dif1 for i in numeromagico] and x!=0 and y!=0:
                                    fechita4=datetime(año,mess,dia,int(hora[0]),int(hora[1]),int(a))
                                    #print("fechita4:",fechita4)
                            hora=hora[0]+":"+hora[1]+":"+a
                            fechas.append(fecha+" - "+hora)
                        elif r==2:
                            datos.append(round(float(row.value),1))
                nlinea=nlinea+1
            delta=fechita2-fechita1
            dif1=delta.days*60*60*24+delta.seconds
            #print("dif1 es:",dif1)
            if dif1>=60*60*24:
                print("faltan más de 24 horas de datos, ver la fila",x)
            delta=fechita4-fechita3
            dif2=delta.days*60*60*24+delta.seconds
            #print("dif2 es:", dif2)
            if dif2>=3600*24:
                print("faltan más de 24 horas de datos, ver la fila",y)
            tituloo=archivo.split()
            maquina=tituloo[3].split(".")[0]
            titulo=mes+"-"+tituloo[2]+"-"+maquina+"-"+tituloo[0]+" "+tituloo[1]
            libro=Workbook()
            hojas=libro.sheetnames
            hojaa=libro[hojas[0]]
            veri=0
            if numeromagico!=hoja.max_row:
                for i in range(len(datos)):
                    k=i+1
                    if k<x+1:
                        posfecha="A"+str(k)
                        posdatos="B"+str(k)
                        hojaa[posfecha]=fechas[i]
                        hojaa[posdatos]=datos[i]
                    elif k==x+1:
                        fechitaa=fechita1+timedelta(0,1)
                        for j in range(dif1):
                            posfecha="A"+str(k+j)
                            posdatos="B"+str(k+j)
                            hojaa[posdatos]=randrange(40,70)########### Aquí se cambia aleatoriedad
                            fechita=str(fechitaa)
                            fechita=fechita.split()
                            hojaa[posfecha]= fechita[0]+" - "+fechita[1]
                            fechitaa+=timedelta(0,1)
                        fechitaa-=timedelta(0,1)
                        hojaa["B"+str(i+dif1)]=datos[i]
                        hojaa["A"+str(i+dif1)]=str(fechitaa).split()[0]+" - "+str(fechitaa).split()[1]
                    elif k>x+1 and k<y+1:
                        posfecha="A"+str(i+dif1)
                        posdatos="B"+str(i+dif1)
                        hojaa[posfecha]=fechas[i]
                        hojaa[posdatos]=datos[i]
                    elif k==y+1:
                        fechitaa=fechita3+timedelta(0,1)
                        for j in range(dif2-1):
                            posfecha="A"+str(k+j+dif1-1)
                            posdatos="B"+str(k+j+dif1-1)
                            hojaa[posdatos]=randrange(40,70)########### Aquí se cambia aleatoriedad
                            fechita=str(fechitaa)
                            fechita=fechita.split()
                            hojaa[posfecha]= fechita[0]+" - "+fechita[1]
                            fechitaa+=timedelta(0,1)
                        if veri==0:
                            #print(fechitaa)
                            #print("A"+str(i+dif1+dif2-1),datos[i])
                            veri=1
                        hojaa["B"+str(i+dif1+dif2-1)]=datos[i]
                        hojaa["A"+str(i+dif1+dif2-1)]=str(fechitaa).split()[0]+" - "+str(fechitaa).split()[1]
                    elif k>y+1:
                        posfecha="A"+str(i+dif1+dif2-1)
                        posdatos="B"+str(i+dif1+dif2-1)
                        hojaa[posfecha]=fechas[i]
                        hojaa[posdatos]=datos[i]

        else:
            for i in hoja.iter_rows():
                for r,row in enumerate(i):
                    if row.value!=None:
                        if r==0:
                            celda=row.value
                            celda=celda.split()
                            fecha=celda[0]
                            fechaa=fecha.split("/")
                            año=int(fechaa[0])
                            mess=int(fechaa[1])
                            dia=int(fechaa[2])
                            hora=celda[2]
                            hora=hora.split(":")
                            segundos=hora[2]
                            a=""
                            contadorS=0
                            for i in segundos:
                                if contadorS<2:
                                    a+=i
                                contadorS+=1
                            hora=hora[0]+":"+hora[1]+":"+a
                            fechas.append(fecha+" - "+hora)
                        elif r==2:
                            datos.append(round(float(row.value),1))
                nlinea+=1
            tituloo=archivo.split()
            maquina=tituloo[3].split(".")[0]
            titulo=mes+"-"+tituloo[2]+"-"+maquina+"-"+tituloo[0]+" "+tituloo[1]
            libro=Workbook()
            hojas=libro.sheetnames
            hojaa=libro[hojas[0]]
            for i in range(len(datos)):
                posfecha="A"+str(i+1)
                posdatos="B"+str(i+1)
                hojaa[posfecha]=fechas[i]
                hojaa[posdatos]=datos[i]                 
              
        

                
                    
                    


        hojaa.column_dimensions["A"].width=23
            
        nombre_archivo=titulo+".xlsx"
        libro.save(filename=nombre_archivo)
    




def semana_del_mes(fecha_inicio,fecha_termino):
    if fecha_inicio.day==1 and fecha_termino.day==7:
        x="1"
    elif fecha_inicio.day==8 and fecha_termino.day==14:
        x="2"
    elif fecha_inicio.day==15 and fecha_termino.day==21:
        x="3"
    elif fecha_inicio.day==22 and fecha_termino.day==28:
        x="4"
    elif fecha_inicio.day==29 and fecha_termino.day<=31:
        x="5"
    return(x)

    
def crear_carpetas_brutas():#voy a ejecutar primero esta funcion y despues la de crear_carpetas_limpias
    años=os.listdir(".")
    os.mkdir("archivos brutos")
    años.remove("ejecutable.py")
    años.remove("instrucciones.pdf")
    febrero=["febrero","Febrero"]
    for i in años:
        os.mkdir("archivos brutos\\"+i)
        for j in os.listdir(".\\"+i):
            os.mkdir("archivos brutos\\"+i+"\\"+j)
            if j not in febrero:
                for k in range(5):
                    os.mkdir("archivos brutos\\"+i+"\\"+j+"\\"+"semana "+str(k+1))
            else:
                for k in range(4):
                    os.mkdir("archivos brutos\\"+i+"\\"+j+"\\"+"semana "+str(k+1))

def crear_carpetas_limpias():
    años=os.listdir(".")
    os.mkdir("archivos limpios")
    años.remove("ejecutable.py")
    años.remove("instrucciones.pdf")
    años.remove("archivos brutos")
    febrero=["febrero","Febrero"]
    for i in años:
        os.mkdir("archivos limpios\\"+i)
        for j in os.listdir(".\\"+i):
            os.mkdir("archivos limpios\\"+i+"\\"+j)
            if j not in febrero:
                for k in range(5):
                    os.mkdir("archivos limpios\\"+i+"\\"+j+"\\"+"semana "+str(k+1))
            else:
                for k in range(4):
                    os.mkdir("archivos limpios\\"+i+"\\"+j+"\\"+"semana "+str(k+1))


def limpiar_uno():
    lugar=puntoV.get()
    maquina=modeloV.get()
    lugar=lugar.upper()
    maquina=maquina.upper()
    fi=fiV.get()
    fi=fi.split()
    
    fechai=fi[0].split("/")
    añoi=int(fechai[0])
    mesi=int(fechai[1])
    diai=int(fechai[2])
    horai=fi[2].split(":")
    horaii=int(horai[0])
    minutoi=int(horai[1])
    segundoi=int(horai[2])
    fecha_inicio=datetime(añoi,mesi,diai,horaii,minutoi,segundoi)
    
    ft=ftV.get()
    ft=ft.split()
    
    fechat=ft[0].split("/")
    añot=int(fechat[0])
    mest=int(fechat[1])
    diat=int(fechat[2])
    horat=ft[2].split(":")
    horatt=int(horat[0])
    
    minutot=int(horat[1])
    segundot=int(horat[2])
    fecha_termino=datetime(añot,mest,diat,horatt,minutot,segundot)
    
    n=semana_del_mes(fecha_inicio,fecha_termino)
    mes=meses[mesi]
    año=str(añoi)

    
    
    numeromagico=crear_archivos_brutos(lugar,maquina,fecha_inicio,fecha_termino,mes,año)
    nombre_archivo_bruto="semana "+n+" "+lugar+" "+maquina+" "+meses[fecha_inicio.month]+" "+str(fecha_inicio.year)+".xlsx"
    limpiar_archivo(nombre_archivo_bruto,mes,numeromagico,fecha_inicio,fecha_termino)
    print("archivos limpio y brutos creados")
    ventana_fin()



def ventana_fin():
    ventana2=Tk()
    ventana2.title("Proceso finalizado")
    etiqueta=Label(ventana2,text="Los archivos han sido analizados y limpiados").grid(row=0,column=0)
    ejecutar=Button(ventana2, text="Cerrar", command=ventana2.destroy).grid(row=1,column=0)
    ventana2.mainloop()


def ventana_instrucciones():
    ventana3=Tk()
    ventana3.title("Instrucciones")
    etiqueta1=Label(ventana3,text="Para obtener cada semana se debe hacer lo siguiente:").place(x=0,y=0)
    etiqueta2=Label(ventana3,text="semana 1:").place(x=0,y=20)
    etiqueta3=Label(ventana3,text="fecha inicio: dia=01, hora=00, minuto=00, segundo=00").place(x=60,y=20)
    etiqueta4=Label(ventana3,text="fecha término: dia=07, hora=23, minuto=59, segundo=59").place(x=60,y=35)
    etiqueta5=Label(ventana3,text="semana 2:").place(x=0,y=60)
    etiqueta6=Label(ventana3,text="fecha inicio: dia=08, hora=00, minuto=00, segundo=00").place(x=60,y=60)
    etiqueta7=Label(ventana3,text="fecha término: dia=14, hora=23, minuto=59, segundo=59").place(x=60,y=75)
    etiqueta8=Label(ventana3,text="semana 3:").place(x=0,y=100)
    etiqueta9=Label(ventana3,text="fecha inicio: dia=15, hora=00, minuto=00, segundo=00").place(x=60,y=100)
    etiqueta10=Label(ventana3,text="fecha término: dia=21, hora=23, minuto=59, segundo=59").place(x=60,y=115)
    etiqueta11=Label(ventana3,text="semana 4:").place(x=0,y=140)
    etiqueta12=Label(ventana3,text="fecha inicio: dia=22, hora=00, minuto=00, segundo=00").place(x=60,y=140)
    etiqueta13=Label(ventana3,text="fecha término: dia=28, hora=23, minuto=59, segundo=59").place(x=60,y=155)
    etiqueta14=Label(ventana3,text="semana 5:").place(x=0,y=180)
    etiqueta15=Label(ventana3,text="fecha inicio: dia=29, hora=00, minuto=00, segundo=00").place(x=60,y=180)
    etiqueta16=Label(ventana3,text="fecha término: dia=31, hora=23, minuto=59, segundo=59").place(x=60,y=195)
    ejecutar2=Button(ventana3, text="Cerrar", command=ventana3.destroy).place(x=175,y=220)
    ventana3.geometry("380x250")


#Ventana:
ventana=Tk()
ventana.title("Filtro de datos")


#Variables:
puntoV=StringVar()
modeloV=StringVar()
fiV=StringVar()
ftV=StringVar()
mesesalimpiarV=StringVar()
semanasalimpiarV=StringVar()
ubicacionesalimpiarV=StringVar()
maquinasalimpiarV=StringVar()

#Etiquetas:
etiqueta=Label(ventana,text="Ingrese datos").grid(row=0,column=0)
punto=Label(ventana,text="Punto:").grid(row=1,column=0)
modelo=Label(ventana,text="Modelo:").grid(row=2,column=0)
fi=Label(ventana,text="fecha inicio (yyyy/mm/dd - hh:mm:ss):").grid(row=3,column=0)
ft=Label(ventana,text="fecha término (yyyy/mm/dd - hh:mm:ss):").grid(row=4,column=0)
mesesalimpiar=Label(ventana,text="meses a limpiar:").grid(row=6,column=0)
semanasalimpiar=Label(ventana,text="semanas a limpiar:").grid(row=7,column=0)
ubicacionesalimpiar=Label(ventana,text="ubicaciones a limpiar:").grid(row=8,column=0)
maquinasalimpiar=Label(ventana,text="maquinas a limpiar:").grid(row=9,column=0)

#Entradas:
puntoVcaja=Entry(ventana,textvariable=puntoV).grid(row=1,column=1)
modeloVcaja=Entry(ventana,textvariable=modeloV).grid(row=2,column=1)
fiVcaja=Entry(ventana,textvariable=fiV).grid(row=3,column=1)
ftVcaja=Entry(ventana,textvariable=ftV).grid(row=4,column=1)
mesesalimpiarVcaja=Entry(ventana,textvariable=mesesalimpiarV).grid(row=6,column=1)
semanasalimpiarVcaja=Entry(ventana,textvariable=semanasalimpiarV).grid(row=7,column=1)
ubicacionesalimpiarVcaja=Entry(ventana,textvariable=ubicacionesalimpiarV).grid(row=8,column=1)
maquinasalimpiarVcaja=Entry(ventana,textvariable=maquinasalimpiarV).grid(row=9,column=1)



#botones:
ejecutar=Button(ventana, text="Limpieza total", command=orden66).grid(row=10,column=0)
limpiar=Button(ventana, text="Limpieza individual", command=limpiar_uno).grid(row=5,column=0)
instrucciones=Button(ventana, text="instrucciones", command=ventana_instrucciones).grid(row=5,column=1)

ventana.mainloop()







